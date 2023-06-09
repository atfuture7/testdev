#######
# xlsxLoc
# a simple tool to find a value in a xlsx file. Since it need to be 
# packaged as a .exe file, it was reversion to a class
# 
# Purpose: 
#   When dealing with shuffled list, finding the occurrance of 
#   a specific value is difficult. This tool is created to assist.
# Condition:
#   It only identify the first occurrance in each column
# 
################

import openpyxl

class xlsxLoc:
    sFileName = "Book.xlsx"
    #wb, ws
    sSearch = ""
    sLoc = []

    # Load file name
    def loadFile(self,sName):
        if len(sName) >0:
            self.sFileName = sName
        self.wb = openpyxl.load_workbook(self.sFileName)
        self.ws = self.wb[self.wb.sheetnames[0]]
        # wb.sheetnames
        # ws.max_row
        # ws.max_column

    # search value in columns. Only the first occurrence
    def cellSearch(self, sTarget):
        self.sLoc.clear()
        self.sLoc = []
        self.sSearch = sTarget
        for oRow in self.ws.columns:
            for cell in oRow:
                if (cell.value == self.sSearch):
                    self.sLoc.append(cell.coordinate)
                    break

    # return search result
    def getResult(self):
        return self.sLoc
        

# if run independently
if __name__ == "__main__":
    import sys
    
    ob = xlsxLoc()
    if len(sys.argv) == 3:
        ob.loadFile(sys.argv[2])  
    else:
        ob.loadFile("")
    ob.cellSearch(sys.argv[1])
    print(ob.getResult())





        